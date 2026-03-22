# -*- coding: utf-8 -*-
import os
from datetime import datetime
from aiogram import Router
from aiogram import F
from aiogram.fsm.context import FSMContext
from aiogram.types import Message, FSInputFile
from loguru import logger  # https://github.com/Delgan/loguru
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from database.database import User, create_keywords_model, get_user_channel_usernames
from locales.locales import get_text

# from system.dispatcher import router

router = Router(name=__name__)


def create_excel_file(data: list, headers: list, filename: str, sheet_name: str) -> str:
    """
    Создаёт Excel-файл с заданными данными, оформлением и автоматической подгонкой ширины столбцов.

    Функция формирует .xlsx файл с синей шапкой заголовков, центрированным жирным белым шрифтом,
    подгоняет ширину столбцов под содержимое и сохраняет файл в директорию 'exports/'.

    Используется для экспорта данных пользователей (ключевые слова, ссылки).

    :param data: (list) Список кортежей, представляющих строки таблицы (например, [(1, 'слово'), ...]).
    :param headers: (list) Список строк, представляющих заголовки столбцов (например, ['№', 'Ключевое слово']).
    :param filename: (str) Имя создаваемого файла (например, "keywords_12345_20250420.xlsx").
    :param sheet_name: (str) Название листа в Excel-файле.
    :return: (str) Абсолютный путь к созданному Excel-файлу.

    Raises:
        Exception: Может быть выброшено при ошибках записи файла (например, нет доступа к папке).
            Ожидается обработка вызывающей стороной.

    Notes:
        - Используется библиотека `openpyxl`.
        - Максимальная ширина столбца ограничена 50 символами.
        - Папка 'exports/' создаётся автоматически при необходимости.
    """
    # Создаём новый Excel-файл (рабочую книгу)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name  # Название листа

    # Настройки оформления заголовков таблицы
    header_font = Font(bold=True, color="FFFFFF", size=12)  # Жирный белый шрифт
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Синий фон
    header_alignment = Alignment(horizontal="center", vertical="center")  # Центрирование текста

    # Запись заголовков в первую строку
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Запись данных (начиная со второй строки)
    for row_num, row_data in enumerate(data, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Автоматическая подгонка ширины столбцов под содержимое
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # ограничиваем ширину 50 символами
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Создаём папку exports, если её ещё нет
    exports_dir = "exports"
    os.makedirs(exports_dir, exist_ok=True)

    # Сохраняем файл
    filepath = os.path.join(exports_dir, filename)
    workbook.save(filepath)
    logger.info(f"Excel файл создан: {filepath}")

    return filepath


@router.message(F.text == "🔍 Список ключевых слов")
async def get_keywords_list(message: Message, state: FSMContext):
    """
    Обработчик команды "🔍 Список ключевых слов" для экспорта ключевых слов в Excel.

    Извлекает все сохранённые ключевые слова пользователя из его персональной таблицы в базе данных,
    формирует Excel-файл с помощью `create_excel_file`, отправляет его пользователю через Telegram
    в виде документа и удаляет файл с сервера после отправки. В случае отсутствия данных или ошибки
    отправляет соответствующее текстовое уведомление.

    - Имя файла генерируется с временной меткой для уникальности.
    - Данные логируются для аудита.
    - Используется динамическая модель `create_keywords_model`.

    :param message: (Message) Входящее сообщение от пользователя, инициировавшего экспорт.
    :param state: (FSMContext) Контекст машины состояний, сбрасывается в начале обработки.
    :return: None
    :raise Exception: Перехватывается локально при ошибках создания или отправки файла.
                      Логируется и преобразуется в пользовательское сообщение об ошибке.
    """
    await state.clear()  # Завершаем текущее состояние машины состояния
    telegram_user = message.from_user
    user = User.get(User.user_id == telegram_user.id)

    logger.info(f"Пользователь {telegram_user.id} {telegram_user.username} запросил экспорт ключевых слов")

    # Получаем модель таблицы ключевых слов для данного пользователя
    KeywordsModel = create_keywords_model(user_id=telegram_user.id)

    # Проверяем, существует ли таблица
    if not KeywordsModel.table_exists():
        KeywordsModel.create_table()
        await message.answer(get_text(user.language, "no_keywords"))
        return

    # Извлекаем все ключевые слова
    keywords = list(KeywordsModel.select())

    if not keywords:
        await message.answer(get_text(user.language, "no_keywords"))
        return

    # Формируем список данных для записи в Excel
    data = []
    for idx, keyword in enumerate(keywords, start=1):
        data.append((idx, keyword.user_keyword))  # Номер и текст ключевого слова

    # Формируем имя файла
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"keywords_{telegram_user.id}_{timestamp}.xlsx"
    headers = ["№", "Ключевое слово / Keyword"]

    try:
        # Создаём Excel-файл
        filepath = create_excel_file(
            data=data,
            headers=headers,
            filename=filename,
            sheet_name="Keywords"
        )

        # Отправляем файл пользователю в Telegram
        document = FSInputFile(filepath)
        await message.answer_document(
            document=document,
            caption=f"📋 {get_text(user.language, 'keywords_export')}\n"
                    f"Всего записей: {len(data)}"
        )

        # Удаляем файл после отправки
        os.remove(filepath)
        logger.info(f"Файл ключевых слов отправлен и удалён: {filepath}")

    except Exception as e:
        logger.exception(f"Ошибка при создании Excel-файла с ключевыми словами: {e}")
        await message.answer(get_text(user.language, "export_error"))


@router.message(F.text == "🌐 Ссылки для отслеживания")
async def get_tracking_links_list(message: Message, state: FSMContext):
    """
    Обработчик команды "🌐 Ссылки для отслеживания" для экспорта списка отслеживаемых чатов.

    Извлекает все сохранённые ссылки (username) на группы и каналы из персональной таблицы пользователя,
    создает Excel-файл с данными, отправляет его пользователю через Telegram и удаляет файл с сервера.
    Если данных нет или произошла ошибка, отправляет соответствующее текстовое уведомление.

    :param message: (Message) Входящее сообщение от пользователя, инициировавшего экспорт.
    :param state: (FSMContext) Контекст машины состояний, сбрасывается в начале обработки.
    """
    await state.clear()
    telegram_user = message.from_user
    user = User.get(User.user_id == telegram_user.id)

    logger.info(f"Пользователь {telegram_user.id} {telegram_user.username} запросил экспорт ссылок для отслеживания")

    # Получаем список username и количество
    usernames_list, total_count = get_user_channel_usernames(telegram_user.id)

    if not usernames_list:
        await message.answer(get_text(user.language, "no_tracking_links"))
        return

    # Формируем список данных для Excel
    data = []
    for idx, username in enumerate(usernames_list, start=1):
        data.append((idx, username))

    # Формируем имя файла
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"tracking_links_{telegram_user.id}_{timestamp}.xlsx"
    headers = ["№", "Username канала/группы / Channel/Group Username"]

    try:
        # Создаём Excel-файл
        filepath = create_excel_file(
            data=data,
            headers=headers,
            filename=filename,
            sheet_name="Tracking Links"
        )

        # Отправляем файл пользователю
        document = FSInputFile(filepath)
        await message.answer_document(
            document=document,
            caption=f"🔗 Всего записей: {total_count}"
        )

        # Удаляем файл после отправки
        os.remove(filepath)
        logger.info(f"Файл ссылок отправлен и удалён: {filepath}")

    except Exception as e:
        logger.exception(f"Ошибка при создании Excel-файла со ссылками: {e}")
        await message.answer(get_text(user.language, "export_error"))

# def register_data_export_handlers():
#     """
#     Регистрирует обработчики для экспорта пользовательских данных в Excel.
#
#     Добавляет в маршрутизатор (router) два обработчика:
#         1. get_keywords_list — для экспорта списка ключевых слов по кнопке "🔍 Список ключевых слов".
#         2. get_tracking_links_list — для экспорта списка отслеживаемых ссылок по кнопке "🌐 Ссылки для отслеживания".
#
#     Эти обработчики позволяют пользователю получать свои данные в виде файлов .xlsx,
#     пригодных для просмотра или анализа в сторонних программах.
#
#     Вызывается при инициализации бота в `main.py`.
#
#     Returns:
#         None
#     """
#     router.message.register(get_keywords_list)
#     router.message.register(get_tracking_links_list)
