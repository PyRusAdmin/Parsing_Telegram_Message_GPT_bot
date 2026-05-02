import asyncio
import io
import re
from datetime import datetime

from aiogram import F, Router
from aiogram.fsm.context import FSMContext
from aiogram.types import BufferedInputFile, ReplyKeyboardRemove, Message
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font
from peewee import IntegrityError
from peewee import fn

from account_manager.auth import CheckingAccountsValidity
from ai.ai import get_groq_response, search_groups_in_telegram
from database.database import User, TelegramGroup
from keyboards.user.keyboards import back_keyboard, search_group_ai, get_categories_keyboard
from locales.locales import t
from states.states import MyStates, ExportStates

router = Router(name=__name__)


def clean_group_name(name):
    """
    Очищает название группы от начальных номеров, символов и лишних пробелов.

    Удаляет с начала строки последовательности из цифр, точек, тире, звёздочек,
    скобок и пробелов, которые часто присутствуют в перечисленных списках.

    Например, преобразует "1. Группа разработчиков" в "Группа разработчиков".

    :param name : (str) Исходное название группы.
    :return str: Очищенное название группы без префиксов.
    """
    cleaned = re.sub(r'^[\d\.\-\*\s\)\(\[\]]+', '', name).strip()
    return cleaned


def save_group_to_db(group_data: dict):
    """
    Сохраняет или обновляет информацию о группе в централизованной базе данных.

    Приоритет проверки:
    1. telegram_id (уникальное поле)
    2. group_hash (fallback, если telegram_id = None)

    При наличии записи — обновляет все поля (включая participants, description и т.д.).
    При отсутствии — создаёт новую запись.

    :param group_data: (dict) Словарь с данными группы
    :return: TelegramGroup or None
    """
    try:
        telegram_id = group_data.get('telegram_id')
        group_hash = group_data.get('group_hash')

        # ========== 1. Проверяем по telegram_id (основной способ) ==========
        if telegram_id is not None:
            existing = TelegramGroup.get_or_none(TelegramGroup.telegram_id == telegram_id)
        else:
            # ========== 2. Fallback — проверяем по group_hash ==========
            existing = TelegramGroup.get_or_none(TelegramGroup.group_hash == group_hash)

        if existing:
            # Обновляем существующую запись
            existing.telegram_id = telegram_id
            existing.group_hash = group_hash
            existing.name = group_data.get('name')
            existing.username = group_data.get('username')
            existing.description = group_data.get('description')
            existing.participants = group_data.get('participants', 0)
            existing.category = (group_data.get('category') or '').lower() or None
            existing.group_type = group_data.get('group_type')
            existing.language = group_data.get('language', '')
            existing.availability = group_data.get('availability', 'unknown')
            existing.link = group_data.get('link')
            # date_added оставляем оригинальным (дата первого добавления)
            # Если хочешь обновлять на "последнее обнаружение" — раскомментируй:

            existing.save()
            logger.info(f"🔄 Обновлена существующая группа: {existing.name} (telegram_id={telegram_id})")
            return existing

        else:
            # Создаём новую запись
            category = group_data.get('category')
            if category:
                category = category.lower()
            new_group = TelegramGroup.create(
                telegram_id=telegram_id,
                group_hash=group_hash,
                name=group_data.get('name'),
                username=group_data.get('username'),
                description=group_data.get('description'),
                participants=group_data.get('participants', 0),
                category=category,
                group_type=group_data.get('group_type'),
                language=group_data.get('language', ''),
                availability=group_data.get('availability', 'unknown'),
                link=group_data.get('link'),
                # date_added автоматически поставится по default в модели
            )
            logger.info(f"✅ Добавлена новая группа: {new_group.name} (telegram_id={telegram_id})")
            return new_group

    except IntegrityError as e:
        if "telegram_groups.telegram_id" in str(e):
            logger.warning(f"Попытка создать дубль по telegram_id. Уже обработано выше.")
            # На всякий случай пробуем обновить ещё раз
            return save_group_to_db(group_data)  # рекурсия 1 раз — безопасно
        else:
            logger.exception(f"Неизвестная IntegrityError при сохранении: {e}")
            return None

    except Exception as e:
        logger.exception(f"Ошибка при сохранении группы: {e}")
        return None


def format_summary_message(groups_count, lang="ru"):
    """
    Форматирует HTML-сообщение с краткой сводкой о результатах поиска.

    Включает статус выполнения, количество найденных групп и уведомление о файле.

    Сообщение отправляется перед XLSX-файлом.

    :param groups_count: (int) Количество успешно сохранённых и отправленных групп.
    :param lang: (str) Язык пользователя
    :return: (str) Сообщение с HTML-разметкой (теги <b>).
    """
    return t(
        "search_summary",
        lang=lang,
        groups_count=groups_count
    )


def create_excel_file(groups):
    """
    Создаёт байтовый Excel-файл (.xlsx) с данными о найденных группах для отправки пользователю.

    Содержит колонки: ID (Hash), Название, Username, Описание, Участников,
    Категория, Тип, Язык, Активность, Ссылка, Дата добавления.
    Username приводится к формату '@username'.

    :param groups: (list[TelegramGroup]) Список экземпляров модели TelegramGroup.
    :return: bytes — содержимое .xlsx файла в памяти.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты поиска"

    # Заголовки
    headers = [
        'ID (Hash)',
        'Название',
        'Username',
        'Описание',
        'Участников',
        'Категория',
        'Тип',
        'Язык',
        'Активность',
        'Ссылка',
        'Дата добавления'
    ]
    ws.append(headers)

    # Жирный шрифт для заголовков
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Данные
    for group in groups:
        username = group.username or ''
        if username:
            username = f"@{username.lstrip('@')}"

        ws.append([
            group.group_hash,
            group.name,
            username,
            group.description or '',
            group.participants,
            group.category or '',
            group.group_type,
            group.language,
            group.availability,
            group.link,
            group.date_added.strftime('%Y-%m-%d %H:%M:%S')
        ])

    # Автоподбор ширины (опционально)
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells) + 2
        ws.column_dimensions[column_cells[0].column_letter].width = min(length, 50)

    # Сохраняем в BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@router.message(F.text == "📥 Вся база")
async def export_all_groups(message: Message, state: FSMContext):
    """Выдаёт CSV-файл со всей базой данных групп и каналов."""
    await state.clear()
    user = User.get(User.user_id == message.from_user.id)

    try:
        # ====================== ОЧИСТКА ДУБЛИКАТОВ ======================
        deleted_count = 0

        # Находим все telegram_id, у которых больше одной записи
        duplicates = (
            TelegramGroup
            .select(
                TelegramGroup.telegram_id,
                fn.COUNT(TelegramGroup.id).alias('cnt')
            )
            .where(TelegramGroup.telegram_id.is_null(False))  # игнорируем NULL
            .group_by(TelegramGroup.telegram_id)
            .having(fn.COUNT(TelegramGroup.id) > 1)
        )

        for dup in duplicates:
            tid = dup.telegram_id

            # Выбираем запись, которую оставляем (самая свежая)
            keep_record = (
                TelegramGroup
                .select(TelegramGroup.id)
                .where(TelegramGroup.telegram_id == tid)
                .order_by(TelegramGroup.date_added.desc())
                .limit(1)
                .get()
            )

            # Удаляем все остальные записи с этим telegram_id
            deleted = (
                TelegramGroup
                .delete()
                .where(
                    (TelegramGroup.telegram_id == tid) &
                    (TelegramGroup.id != keep_record.id)
                )
                .execute()
            )
            deleted_count += deleted

        if deleted_count > 0:
            logger.info(f"✅ Перед экспортом очищено {deleted_count} дубликатов по telegram_id")

        # ====================== ЭКСПОРТ ======================
        groups = TelegramGroup.select()
        if not groups:
            await message.answer(t("database_empty", lang=user.language))
            return

        excel_bytes = create_excel_file(groups)
        document = BufferedInputFile(excel_bytes, filename="Вся_база.xlsx")
        await message.answer_document(
            document=document,
            caption=t(
                "export_all_caption",
                lang=user.language,
                total_records=len(groups),
                deleted_duplicates=deleted_count
            )
        )

    except Exception as e:
        await message.answer(t("export_error_generic", lang=user.language))
        logger.exception(e)


@router.message(F.text == "📥 База каналов")
async def export_channels(message: Message, state: FSMContext):
    """Выдаёт CSV-файл со всей базой данных групп и каналов."""
    await state.clear()
    user = User.get(User.user_id == message.from_user.id)
    try:
        # Получаем только КАНАЛЫ
        groups = TelegramGroup.select().where(
            TelegramGroup.group_type == 'Канал'
        )

        if not groups:
            await message.answer(t("database_empty", lang=user.language))
            return

        excel_bytes = create_excel_file(groups)
        document = BufferedInputFile(excel_bytes, filename="База_каналов.xlsx")
        await message.answer_document(
            document=document,
            caption=t("export_channels_caption", lang=user.language, total_records=len(groups))
        )

    except Exception as e:
        await message.answer(t("export_error_generic", lang=user.language))
        logger.exception(e)


@router.message(F.text == "📥 База групп")
async def export_supergroups(message: Message, state: FSMContext):
    """Выдаёт CSV-файл со всей базой данных групп и каналов."""
    await state.clear()
    user = User.get(User.user_id == message.from_user.id)
    try:
        # Получаем только СУПЕРГРУППЫ
        groups = TelegramGroup.select().where(
            TelegramGroup.group_type == 'Группа (супергруппа)'
        )
        if not groups:
            await message.answer(t("database_empty", lang=user.language))
            return

        excel_bytes = create_excel_file(groups)
        document = BufferedInputFile(excel_bytes, filename="База_групп.xlsx")
        await message.answer_document(
            document=document,
            caption=t("export_groups_caption", lang=user.language, total_records=len(groups))
        )

    except Exception as e:
        await message.answer(t("export_error_generic", lang=user.language))
        logger.exception(e)


@router.message(F.text == "📥 Получить базу")
async def handle_enter_keyword_menu(message: Message, state: FSMContext):
    """
    Обрабатывает запрос пользователя на получение базы Telegram-групп и каналов.

    Отображает информационное сообщение с описанием доступных действий:
    - 📥 Получение всей базы данных
    - 🔙 Возврат в главное меню

    Используется как промежуточное меню для навигации в разделе поиска.

    :param message: (Message) Входящее сообщение от пользователя.
    :param state: (FSMContext, optional) Контекст состояния конечного автомата (не используется, но передаётся).
    :return: None
    """
    await state.clear()
    user = User.get(User.user_id == message.from_user.id)

    await message.answer(
        text=t("get_database_menu", lang=user.language),
        reply_markup=search_group_ai(),
        parse_mode="HTML"
    )


@router.message(F.text == "📂 Выбрать категорию")
async def start_category_export(message: Message, state: FSMContext):
    """
    Запускает процесс выбора категории для экспорта.
    Показывает клавиатуру с категориями и переводит в состояние ожидания выбора.
    """
    user = User.get(User.user_id == message.from_user.id)
    await message.answer(
        t("select_category_prompt", lang=user.language),
        reply_markup=get_categories_keyboard()
    )
    await state.set_state(ExportStates.waiting_for_category)


@router.message(ExportStates.waiting_for_category)
async def handle_category_selection(message: Message, state: FSMContext):
    """
    Обрабатывает выбор категории и формирует файл со списком групп.
    """
    user = User.get(User.user_id == message.from_user.id)
    selected_category = message.text.strip()

    # Проверка на кнопку "Назад"
    if selected_category == "⬅️ Назад":
        await message.answer(t("action_cancelled", lang=user.language), reply_markup=ReplyKeyboardRemove())
        await state.clear()
        return

    # Список допустимых категорий (нижний регистр)
    selected_category = selected_category.lower()
    valid_categories = {
        "инвестиции",
        "финансы и личный бюджет",
        "криптовалюты и блокчейн",
        "бизнес и предпринимательство",
        "маркетинг и продвижение",
        "технологии и it",
        "образование и саморазвитие",
        "работа и карьера",
        "недвижимость",
        "здоровье и медицина",
        "путешествия",
        "авто и транспорт",
        "шоппинг и скидки",
        "развлечения и досуг",
        "политика и общество",
        "наука и исследования",
        "спорт и фитнес",
        "кулинария и еда",
        "мода и красота",
        "хобби и творчество"
    }

    if selected_category not in valid_categories:
        await message.answer(
            t("invalid_category", lang=user.language),
            reply_markup=get_categories_keyboard()
        )
        return

    # Получаем группы из базы
    groups = TelegramGroup.select().where(TelegramGroup.category == selected_category)
    group_count = groups.count()

    if group_count == 0:
        await message.answer(
            t("category_empty", lang=user.language, category=selected_category),
            reply_markup=ReplyKeyboardRemove()
        )
        await state.clear()
        return

    # === Создаём Excel-файл ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Группы"

    # Заголовки
    headers = ["Username", "Название", "Описание", "Тип", "Участники", "Ссылка"]
    ws.append(headers)

    # Жирный шрифт для заголовков
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Данные
    for group in groups:
        ws.append([
            group.username or "",
            group.name or "",
            group.description or "",
            group.group_type or "",
            group.participants or 0,
            group.link or ""
        ])

    # Автоподбор ширины колонок (опционально)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # ограничим ширину
        ws.column_dimensions[column].width = adjusted_width

    # Сохраняем в память
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Отправляем файл
    file_name = f"groups_{selected_category.replace(' ', '_')}.xlsx"
    await message.answer_document(
        document=BufferedInputFile(
            file=output.getvalue(),
            filename=file_name
        ),
        caption=t("category_export_caption", lang=user.language, group_count=group_count, category=selected_category),
        reply_markup=ReplyKeyboardRemove()
    )

    logger.info(f"Пользователь {message.from_user.id} экспортировал Excel по категории: {selected_category}")
    await state.clear()


"""Одиночный AI поиск"""


@router.message(F.text == "🤖 AI поиск")
async def ai_search(message: Message, state: FSMContext):
    """
    Обработчик команды "Получить базу".

    Очищает состояние FSM, получает данные пользователя, логирует действие
    и запрашивает у пользователя ключевое слово для поиска групп через AI.
    Переводит пользователя в состояние ожидания ввода (MyStates.entering_keyword_ai_search).

    :param message: (Message) Входящее сообщение от пользователя.
    :param state: (FSMContext) Контекст машины состояний, сбрасывается при входе.
    :return: None
    """
    await state.clear()  # Сбрасывает состояние

    telegram_user = message.from_user
    user = User.get(User.user_id == telegram_user.id)

    logger.info(
        f"Пользователь {telegram_user.id} {telegram_user.username} перешел в меню поиска групп")

    await message.answer(
        t("enter_keyword", lang=user.language),
        reply_markup=back_keyboard()
    )
    await state.set_state(MyStates.entering_keyword_ai_search)


@router.message(MyStates.entering_keyword_ai_search)
async def handle_enter_keyword(message: Message, state: FSMContext):
    """
    Обработчик ввода ключевого слова для AI-поиска групп и каналов.

    Получает запрос от пользователя, генерирует варианты названий через Groq API,
    ищет соответствующие группы в Telegram, сохраняет их в базу данных и отправляет
    результаты пользователю в виде XLSX-файла.

    В процессе показывает статус "Ищу...", удаляет его после завершения и отправляет
    сводку и файл.

    Обрабатывает ошибки и пустые результаты.

    - Использует `get_groq_response` для генерации названий.
    - Использует `search_groups_in_telegram` для поиска в Telegram.
    - Результаты сохраняются через `save_group_to_db`.
    - Файл создаётся через `create_excel_file` и отправляется как документ.

    :param message: (Message) Входящее сообщение с ключевым словом.
    :param state: (FSMContext) Контекст машины состояний, сбрасывается после обработки.
    :return: None

    Raises:
        Exception: Перехватывается локально, логируется и преобразуется в пользовательское сообщение.
    """
    # telegram_user = message.from_user
    user_input = message.text.strip()
    user = User.get(User.user_id == message.from_user.id)
    # Отправляем сообщение о начале поиска
    processing_msg = await message.answer(t("searching_groups", lang=user.language))

    try:
        answer = await get_groq_response(user_input)  # Получаем ответ от AI
        logger.info(f"Ответ от Groq: {answer}")

        # Разбиваем ответ на строки и очищаем
        group_names = [clean_group_name(line) for line in answer.splitlines() if line.strip()]
        group_names = [name for name in group_names if len(name) > 2]
        logger.info(f"Получено {len(group_names)} названий: {group_names}")

        saved_groups = []

        # ✅ Создаем checker БЕЗ path (он не нужен для работы с БД)
        checker = CheckingAccountsValidity(message=message)  # path=None по умолчанию
        # client = None
        try:
            client = await checker.start_random_client()
        except Exception as e:
            logger.error(f"❌ Ошибка запуска клиента: {e}")
            await state.clear()
            return

        if not client:
            await message.answer(
                t("search_no_available_accounts", lang=user.language),
                reply_markup=back_keyboard()
            )
            await state.clear()
            return

        for group_name in group_names:
            # Ищем в Telegram
            results = await search_groups_in_telegram(
                client=client,
                group_names=[group_name]
            )
            logger.info(f"Найдено {len(results)} групп для '{group_name}'")

            # Сохраняем результаты в БД
            for group_data in results:
                saved_group = save_group_to_db(group_data)
                if saved_group:
                    saved_groups.append(saved_group)

        # Удаляем сообщение о поиске
        await processing_msg.delete()

        # Отправляем результаты пользователю
        if saved_groups:

            # Создаём Excel-файл
            excel_bytes = create_excel_file(saved_groups)
            filename = f"telegram_groups_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            excel_file = BufferedInputFile(excel_bytes, filename=filename)

            summary = format_summary_message(len(saved_groups), lang=user.language)
            await message.answer(summary, parse_mode="HTML")
            # Отправляем CSV файл
            await message.answer_document(
                document=excel_file,
                caption=t("search_results_caption", lang=user.language, query=user_input),
                parse_mode="HTML"
            )
            logger.info(f"Отправлено {len(saved_groups)} групп пользователю {message.from_user.id} в Excel файле")
        else:
            await message.answer(
                t("search_no_results", lang=user.language),
                reply_markup=back_keyboard()
            )
    except Exception as e:
        logger.error(f"Ошибка при обработке запроса: {e}")
        await processing_msg.delete()
        await message.answer(
            t("search_error", lang=user.language),
            reply_markup=back_keyboard()
        )
    finally:
        if client:
            await client.disconnect()
        await state.clear()  # Завершаем текущее состояние машины состояния


"""Глобальный AI поиск"""


@router.message(F.text == "🌐 Глобальный AI поиск")
async def ai_search_global(message: Message, state: FSMContext):
    """
    Обработчик команды "Глобальный AI поиск".
    Запрашивает у пользователя ключевое слово (или список) для поиска.
    """
    await state.clear()

    telegram_user = message.from_user
    user = User.get(User.user_id == telegram_user.id)

    logger.info(
        f"Пользователь {telegram_user.id} {telegram_user.username} перешел в меню глобального поиска групп"
    )

    await message.answer(
        t("enter_keyword", lang=user.language),
        reply_markup=back_keyboard()
    )
    await state.set_state(MyStates.entering_keyword_ai_search_global)


@router.message(MyStates.entering_keyword_ai_search_global)
async def handle_enter_keyword(message: Message, state: FSMContext):
    """
    Обработчик ввода ключевого слова (или списка) для AI-поиска.
    Каждый запрос обрабатывается через ОТДЕЛЬНЫЙ случайный аккаунт.
    """
    telegram_user = message.from_user
    user = User.get(User.user_id == telegram_user.id)
    user_input = message.text.strip()

    # Парсим ввод в список запросов
    search_terms = parse_search_input(user_input)

    if not search_terms:
        await message.answer(
            t("global_search_no_terms", lang=user.language),
            reply_markup=back_keyboard()
        )
        await state.clear()
        return

    processing_msg = await message.answer(t("global_search_processing", lang=user.language, total=len(search_terms)))

    all_saved_groups = []
    successful_queries = 0

    try:
        # 🔄 Обрабатываем КАЖДЫЙ запрос через НОВЫЙ случайный аккаунт
        for idx, term in enumerate(search_terms, 1):
            logger.info(f"[{idx}/{len(search_terms)}] Запрос: '{term}'")

            # ✅ Создаем checker БЕЗ path (он не нужен для работы с БД)
            checker = CheckingAccountsValidity(message=message)  # path=None по умолчанию
            client = None
            try:
                client = await checker.start_random_client()
            except Exception as e:
                logger.error(f"❌ Ошибка запуска клиента для '{term}': {e}")
                continue

            if not client:
                logger.warning(f"⚠️ Не удалось запустить клиент для '{term}', пропускаю")
                await message.answer(t("global_search_skipped", lang=user.language, term=term))
                continue

            try:
                # Получаем варианты названий от AI
                answer = await get_groq_response(term)
                logger.info(f"Ответ от Groq для '{term}': {answer}")

                # Чистим и фильтруем названия
                group_names = [
                    clean_group_name(line)
                    for line in answer.splitlines()
                    if line.strip() and len(clean_group_name(line)) > 2
                ]

                if not group_names:
                    logger.info(f"⚪ Нет названий для '{term}' после очистки")
                    continue

                logger.info(f"🔍 Ищу {len(group_names)} вариантов для '{term}'")

                # Ищем группы в Telegram
                results = await search_groups_in_telegram(
                    client=client,
                    group_names=group_names
                )
                logger.info(f"✅ Найдено {len(results)} групп для '{term}'")

                # Сохраняем в БД
                for group_data in results:
                    saved_group = save_group_to_db(group_data)
                    if saved_group:
                        all_saved_groups.append(saved_group)

                successful_queries += 1

                # 📊 Обновляем статус в Telegram (опционально)
                if idx % 3 == 0 or idx == len(search_terms):  # каждые 3 запроса или в конце
                    await processing_msg.edit_text(
                        t("global_search_progress", lang=user.language, current=idx, total=len(search_terms),
                          successful=successful_queries)
                    )

            except Exception as e:
                logger.warning(f"⚠️ Ошибка при обработке '{term}': {e}")
                continue  # Продолжаем со следующим запросом

            finally:
                # 🔌 Обязательно отключаем клиент после каждого запроса
                if client:
                    await client.disconnect()
                    logger.info(f"🔌 Клиент для '{term}' отключён")

                # Пауза между запросами (защита от лимитов API и Telegram)
                if idx < len(search_terms):
                    await asyncio.sleep(2)

        await processing_msg.delete()

        # 📤 Отправляем результаты
        if all_saved_groups:
            excel_bytes = create_excel_file(all_saved_groups)
            filename = f"telegram_groups_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            excel_file = BufferedInputFile(excel_bytes, filename=filename)

            summary = format_summary_message(len(all_saved_groups), lang=user.language)
            await message.answer(summary, parse_mode="HTML")

            await message.answer_document(
                document=excel_file,
                caption=t("global_search_results_caption", lang=user.language, total=len(all_saved_groups),
                          successful=successful_queries, total_queries=len(search_terms)),
                parse_mode="HTML"
            )
            logger.info(f"✅ Отправлено {len(all_saved_groups)} групп пользователю {telegram_user.id}")
        else:
            await message.answer(
                t("global_search_no_results", lang=user.language),
                reply_markup=back_keyboard()
            )

    except Exception as e:
        logger.error(f"❌ Критическая ошибка: {e}")
        await processing_msg.delete()
        await message.answer(
            t("search_error", lang=user.language),
            reply_markup=back_keyboard()
        )
    finally:
        await state.clear()


def parse_search_input(user_input: str) -> list[str]:
    """
    Преобразует пользовательский ввод в список поисковых запросов.
    Поддерживает разделители: \n, \r\n, ',', ';'
    Убирает пустые строки и дубликаты, сохраняя порядок.
    """
    if not user_input or not user_input.strip():
        return []

    # Нормализуем разделители → перенос строки
    normalized = user_input.replace(',', '\n').replace(';', '\n').replace('\r\n', '\n')

    # Чистим, фильтруем пустые, убираем дубликаты с сохранением порядка
    seen = set()
    result = []
    for line in normalized.splitlines():
        cleaned = line.strip()
        if cleaned and cleaned not in seen:
            result.append(cleaned)
            seen.add(cleaned)

    return result
